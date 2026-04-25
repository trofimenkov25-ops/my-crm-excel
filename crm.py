import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class ExcelCRM:
    def __init__(self, filename='CRM_Database.xlsx'):
        self.filename = filename
        self.workbook = None
        self.worksheet = None
        self.init_excel()
    
    def init_excel(self):
        """Инициализировать или загрузить Excel файл"""
        if os.path.exists(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
        else:
            self.create_new_file()
    
    def create_new_file(self):
        """Создать новый Excel файл с заголовками"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Контакты"
        
        # Заголовки колонок
        headers = [
            'ID',
            'Категория',
            'Вид деятельности',
            'Название компании',
            'ФИО',
            'Должность',
            'Телефон',
            'Email',
            'WhatsApp',
            'Telegram',
            'Instagram',
            'Соцсети',
            'Сайт',
            'ИНН ЮЛ',
            'ЛПР',
            'Примечание',
            'Дата создания',
            'Дата обновления'
        ]
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Установить ширину колонок
        column_widths = {
            'A': 5, 'B': 12, 'C': 18, 'D': 20, 'E': 18, 'F': 15,
            'G': 15, 'H': 20, 'I': 15, 'J': 15, 'K': 15, 'L': 20,
            'M': 20, 'N': 12, 'O': 15, 'P': 25, 'Q': 18, 'R': 18
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        # Заморозить заголовки
        self.worksheet.freeze_panes = 'A2'
        
        self.save()
    
    def add_contact(self, category, activity_type, company_name, full_name, position,
                   phone, email, whatsapp, telegram, instagram, social_networks,
                   website, inn, lpr, notes):
        """Добавить новый контакт"""
        next_row = self.worksheet.max_row + 1
        contact_id = next_row - 1
        
        data = [
            contact_id,
            category,
            activity_type,
            company_name,
            full_name,
            position,
            phone,
            email,
            whatsapp,
            telegram,
            instagram,
            social_networks,
            website,
            inn,
            lpr,
            notes,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, value in enumerate(data, 1):
            cell = self.worksheet.cell(row=next_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self.save()
        print(f"✅ Контакт добавлен (ID: {contact_id})")
    
    def search_contacts(self, search_term):
        """Поиск контактов по названию, имени или email"""
        results = []
        search_lower = search_term.lower()
        
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            
            # Поиск по компании, имени, email, телефону
            if (str(row[3]).lower() if row[3] else '').find(search_lower) >= 0 or \
               (str(row[4]).lower() if row[4] else '').find(search_lower) >= 0 or \
               (str(row[7]).lower() if row[7] else '').find(search_lower) >= 0 or \
               (str(row[6]).lower() if row[6] else '').find(search_lower) >= 0:
                results.append(row)
        
        return results
    
    def filter_by_category(self, category):
        """Фильтр по категории"""
        results = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if row[1] == category:
                results.append(row)
        return results
    
    def filter_by_activity(self, activity):
        """Фильтр по виду деятельности"""
        results = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if row[2] == activity:
                results.append(row)
        return results
    
    def get_all_contacts(self):
        """Получить все контакты"""
        contacts = []
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            contacts.append(row)
        return contacts
    
    def get_statistics(self):
        """Получить статистику"""
        stats = {
            'total_contacts': 0,
            'categories': {},
            'activities': {}
        }
        
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            
            stats['total_contacts'] += 1
            
            # Подсчёт по категориям
            category = row[1]
            if category:
                stats['categories'][category] = stats['categories'].get(category, 0) + 1
            
            # Подсчёт по видам деятельности
            activity = row[2]
            if activity:
                stats['activities'][activity] = stats['activities'].get(activity, 0) + 1
        
        return stats
    
    def update_contact(self, contact_id, **kwargs):
        """Обновить контакт"""
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2), start=2):
            if row[0].value == contact_id:
                for col_name, value in kwargs.items():
                    col_map = {
                        'category': 2, 'activity_type': 3, 'company_name': 4,
                        'full_name': 5, 'position': 6, 'phone': 7, 'email': 8,
                        'whatsapp': 9, 'telegram': 10, 'instagram': 11,
                        'social_networks': 12, 'website': 13, 'inn': 14,
                        'lpr': 15, 'notes': 16
                    }
                    if col_name in col_map:
                        self.worksheet.cell(row=row_idx, column=col_map[col_name]).value = value
                
                # Обновить дату последнего изменения
                self.worksheet.cell(row=row_idx, column=18).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save()
                print(f"✅ Контакт {contact_id} обновлён")
                return True
        
        print(f"❌ Контакт {contact_id} не найден")
        return False
    
    def delete_contact(self, contact_id):
        """Удалить контакт"""
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2), start=2):
            if row[0].value == contact_id:
                self.worksheet.delete_rows(row_idx, 1)
                self.save()
                print(f"✅ Контакт {contact_id} удалён")
                return True
        
        print(f"❌ Контакт {contact_id} не найден")
        return False
    
    def export_to_csv(self, filename='export.csv'):
        """Экспортировать в CSV"""
        import csv
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in self.worksheet.iter_rows(values_only=True):
                writer.writerow(row)
        print(f"✅ Экспортировано в {filename}")
    
    def save(self):
        """Сохранить файл"""
        self.workbook.save(self.filename)

# ============ ИНТЕРАКТИВНОЕ МЕНЮ ============

def print_menu():
    print("\n" + "="*60)
    print("📊 CRM СИСТЕМА - УПРАВЛЕНИЕ КОНТАКТАМИ В EXCEL")
    print("="*60)
    print("1️⃣  Добавить новый контакт")
    print("2️⃣  Поиск контакта")
    print("3️⃣  Фильтр по категории")
    print("4️⃣  Фильтр по виду деятельности")
    print("5️⃣  Просмотреть все контакты")
    print("6️⃣  Редактировать контакт")
    print("7️⃣  Удалить контакт")
    print("8️⃣  Статистика")
    print("9️⃣  Экспортировать в CSV")
    print("0️⃣  Выход")
    print("="*60)

def display_contacts(contacts, title=""):
    if not contacts:
        print("❌ Контакты не найдены")
        return
    
    if title:
        print(f"\n{title}")
    print("-" * 150)
    print(f"{'ID':<4} {'Категория':<12} {'Вид дейст.':<15} {'Компания':<20} {'ФИО':<18} {'Телефон':<15} {'Email':<20}")
    print("-" * 150)
    
    for contact in contacts:
        print(f"{contact[0]:<4} {str(contact[1]):<12} {str(contact[2]):<15} {str(contact[3]):<20} {str(contact[4]):<18} {str(contact[6]):<15} {str(contact[7]):<20}")
    
    print("-" * 150)

def main():
    crm = ExcelCRM('CRM_Database.xlsx')
    
    while True:
        print_menu()
        choice = input("Выберите действие (0-9): ").strip()
        
        if choice == '1':
            print("\n➕ ДОБАВИТЬ НОВЫЙ КОНТАКТ")
            print("-" * 40)
            category = input("Категория (Партнёр/Клиент/Поставщик/Конкурент): ").strip()
            activity = input("Вид деятельности: ").strip()
            company = input("Название компании: ").strip()
            name = input("ФИО: ").strip()
            position = input("Должность: ").strip()
            phone = input("Телефон: ").strip()
            email = input("Email: ").strip()
            whatsapp = input("WhatsApp (номер): ").strip()
            telegram = input("Telegram (@username): ").strip()
            instagram = input("Instagram (@username): ").strip()
            social = input("Соцсети (ссылки): ").strip()
            website = input("Сайт: ").strip()
            inn = input("ИНН ЮЛ: ").strip()
            lpr = input("ЛПР (Лицо, принимающее решение): ").strip()
            notes = input("Примечание: ").strip()
            
            crm.add_contact(category, activity, company, name, position, phone,
                          email, whatsapp, telegram, instagram, social, website, inn, lpr, notes)
        
        elif choice == '2':
            search_term = input("\n🔍 Введите поисковый запрос (имя, компания, email): ").strip()
            results = crm.search_contacts(search_term)
            display_contacts(results, f"📌 Найдено контактов: {len(results)}")
        
        elif choice == '3':
            category = input("\n🏷️  Выберите категорию (Партнёр/Клиент/Поставщик/Конкурент): ").strip()
            results = crm.filter_by_category(category)
            display_contacts(results, f"📌 Контакты категории '{category}': {len(results)}")
        
        elif choice == '4':
            activity = input("\n📋 Введите вид деятельности: ").strip()
            results = crm.filter_by_activity(activity)
            display_contacts(results, f"📌 Контакты вида '{activity}': {len(results)}")
        
        elif choice == '5':
            contacts = crm.get_all_contacts()
            display_contacts(contacts, f"📌 Всего контактов: {len(contacts)}")
        
        elif choice == '6':
            contact_id = int(input("\n��️  Введите ID контакта для редактирования: ").strip())
            print("Введите новые данные (оставьте пустым, чтобы не менять):")
            
            updates = {}
            fields = ['category', 'activity_type', 'company_name', 'full_name', 'position',
                     'phone', 'email', 'whatsapp', 'telegram', 'instagram', 'social_networks',
                     'website', 'inn', 'lpr', 'notes']
            labels = ['Категория', 'Вид деятельности', 'Компания', 'ФИО', 'Должность',
                     'Телефон', 'Email', 'WhatsApp', 'Telegram', 'Instagram', 'Соцсети',
                     'Сайт', 'ИНН', 'ЛПР', 'Примечание']
            
            for field, label in zip(fields, labels):
                value = input(f"{label}: ").strip()
                if value:
                    updates[field] = value
            
            if updates:
                crm.update_contact(contact_id, **updates)
            else:
                print("❌ Ничего не изменено")
        
        elif choice == '7':
            contact_id = int(input("\n🗑️  Введите ID контакта для удаления: ").strip())
            confirm = input("Вы уверены? (да/нет): ").strip().lower()
            if confirm == 'да':
                crm.delete_contact(contact_id)
        
        elif choice == '8':
            stats = crm.get_statistics()
            print("\n📊 СТАТИСТИКА")
            print("-" * 40)
            print(f"Всего контактов: {stats['total_contacts']}")
            print("\nПо категориям:")
            for cat, count in stats['categories'].items():
                print(f"  • {cat}: {count}")
            print("\nПо видам деятельности:")
            for act, count in stats['activities'].items():
                print(f"  • {act}: {count}")
        
        elif choice == '9':
            filename = input("\n📥 Введите имя файла для экспорта (по умолчанию export.csv): ").strip()
            if not filename:
                filename = 'export.csv'
            crm.export_to_csv(filename)
        
        elif choice == '0':
            print("\n👋 До свидания!")
            break
        
        else:
            print("❌ Неверный выбор. Попробуйте снова.")

if __name__ == '__main__':
    main()